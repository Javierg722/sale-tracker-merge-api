from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
import json
import traceback

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

SHEET_NAME = "1_Data Entry"
START_ROW = 6
END_ROW = 505

INPUT_COLUMNS = {
    "ticker": "E",
    "buyDate": "G",
    "sharesBought": "H",
    "costPerShare": "I",
    "sellDate": "J",
    "sharesSold": "K",
    "salePricePerShare": "L",
    "note": "N",
}

def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value), "%Y-%m-%d")
    except Exception:
        return None

@app.get("/api/merge")
def merge_get():
    return JSONResponse(
        status_code=405,
        content={"detail": "Method not allowed"}
    )

@app.post("/api/merge")
async def merge(
    workbook: UploadFile = File(...),
    data: str = Form(...)
):
    try:
        rows = json.loads(data)
        if not isinstance(rows, list):
            return JSONResponse(
                status_code=400,
                content={"error": "Data must be a JSON array"}
            )

        contents = await workbook.read()
        wb = load_workbook(filename=BytesIO(contents))

        if SHEET_NAME not in wb.sheetnames:
            return JSONResponse(
                status_code=400,
                content={"error": f'Sheet "{SHEET_NAME}" not found'}
            )

        ws = wb[SHEET_NAME]

        # Force recalculation when Excel opens
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
        except Exception:
            pass

        # Clear only the app-owned input cells
        for row_num in range(START_ROW, END_ROW + 1):
            for col in INPUT_COLUMNS.values():
                ws[f"{col}{row_num}"].value = None

        # Write rows back into template
        for i, row in enumerate(rows, start=START_ROW):
            if i > END_ROW:
                return JSONResponse(
                    status_code=400,
                    content={"error": "Too many rows for workbook template"}
                )

            ticker = row.get("ticker")
            buy_date = row.get("buyDate")
            shares_bought = row.get("sharesBought")
            cost_per_share = row.get("costPerShare")
            sell_date = row.get("sellDate")
            shares_sold = row.get("sharesSold")
            sale_price_per_share = row.get("salePricePerShare")
            note = row.get("note")

            if ticker not in (None, ""):
                ws[f"E{i}"] = str(ticker)

            if buy_date not in (None, ""):
                dt = parse_date(buy_date)
                if dt:
                    ws[f"G{i}"] = dt
                    ws[f"G{i}"].number_format = "m/d/yyyy"

            if shares_bought not in (None, ""):
                ws[f"H{i}"] = float(shares_bought)

            if cost_per_share not in (None, ""):
                ws[f"I{i}"] = float(cost_per_share)

            if sell_date not in (None, ""):
                dt = parse_date(sell_date)
                if dt:
                    ws[f"J{i}"] = dt
                    ws[f"J{i}"].number_format = "m/d/yyyy"

            if shares_sold not in (None, ""):
                ws[f"K{i}"] = float(shares_sold)

            if sale_price_per_share not in (None, ""):
                ws[f"L{i}"] = float(sale_price_per_share)

            if note not in (None, ""):
                ws[f"N{i}"] = str(note)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="merged.xlsx"'
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "error": str(e),
                "trace": traceback.format_exc()
            }
        )
